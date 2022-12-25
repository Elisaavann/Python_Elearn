import csv
import itertools
import doctest
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
import matplotlib.pyplot as plt
import numpy as np
from jinja2 import Environment, FileSystemLoader
import pdfkit


# dic_money (dict): Глобальная переменная-словарь. 
# Ключи - кода валют, 
# Значения - словари, содержащие русские наименования валют и их курс
dic_money = {'AZN':{'name':'Манаты','cost':35.68},'BYR':{'name':'Белорусские рубли','cost': 23.91},
    'EUR':{'name':'Евро','cost': 59.90},'GEL':{'name':'Грузинский лари','cost': 21.74},
    'KGS':{'name':'Киргизский сом','cost': 0.76},'KZT':{'name':'Тенге','cost': 0.13},
    'RUR':{'name':'Рубли','cost':1.00},'UAH':{'name':'Гривны','cost':  1.64},
    'USD':{'name':'Доллары','cost': 60.66},'UZS':{'name':'Узбекский сум','cost':0.0055}}

class Salary:
    """Класс для представления зарплаты.

    Attributes: 
        salary_from (str or int or float): Нижняя граница оклада
        salary_to (str or int or float): Верхняя граница оклада
        salary_currency (str): Валюта оклада
    """
    def __init__(self, dict_vac):
        """Инициализирует объект Salary.

        Args: 
            dict_vac (dict): Словарь прочитанных данных для текущей вакансии.
        
        Returns:
            Экземпляр класса Salary
        """
        self.salary_from = dict_vac['salary_from']
        self.salary_to = dict_vac['salary_to']
        self.salary_currency = dict_vac['salary_currency']
    
class Vacancy:
    """Класс для представления данных об одной вакансии.

    Attributes: 
        name (str): Название вакансии
        salary (__main__.Salary): объект Salary, содержит данные о зарплате 
        area_name (str): Город размещения вакансии
        published_at (str): Дата публикации вакансии (строковое представление UTF)
    """
    def __init__(self, dict_vac):
        """Инициализирует объект Vacancy.

        Args: 
            dict_vac (dict): Словарь прочитанных данных для текущей вакансии.
        
        Returns:
            Экземпляр класса Vacancy
        """
        self.name = dict_vac['name']
        self.salary = Salary(dict_vac)
        self.area_name = dict_vac['area_name']
        self.published_at = dict_vac['published_at']


class DataSet:
    """Класс для представления данных о всех вакансиях.

    Attributes:
        file_name (str): Имя файла csv с данными о вакансиях
        vacancies_objects (list[Vacancy]): Лист экземпляров Vacancy хранящий данные о всех прочитанных из file_name вакансиях
        dynamics_objects (__main__.DinamicObjects): Результаты первичной стат. обработки данных о вакансиях 
    """
    @staticmethod    
    def _сsv_reader(filename):
        """Внутренний метод класса. Читает данные о вакансиях из файла filename
        
        Args:
            filename (str): Имя файла с данными о вакансиях

        Returns:
            Список вакансий в виде list[list]. Из списка исключены вакансии, содержащие пустые поля
        """

        result = []
        with open(filename, encoding='utf-8-sig', newline='') as f:
            reader = csv.reader(f)
            for cur_row in reader:
                if ('' in cur_row): continue
                result.append(cur_row)
        return result

    @staticmethod
    def _csv_ﬁler(res_data):
        """Внутренний метод класса. Чистит данные о вакансиях. Преобразует данные в список словарей
        
        Args:
            res_data (list[list]): Список списков данных о вакансиях. Первая запись должна содержать наименования 'свойств' вакансий

        Returns:
            Список вакансий в виде списка словарей (list[dict]). Словарь для каждой вакансии организован как 
            {имя свойства вакансии (str): значение (str)}
        """

        if (len(res_data) == 0): return [{}]
        res_head = res_data.pop(0)
        return [dict(zip(res_head, row)) for row in res_data]
    
    @staticmethod
    def _csv_parser(task):
        """Внутренний метод класса. Вызывает чтение и первичную очистку данных. 
           Удаляет 'грязь' из данных (теги, лишние пробелы и другую служебку).
           В данной версии т.к. по условию входные данные чистые эта часть удалена (добавлять из 5.2 Функции в ООП).
        
        Args:
            task (__main__.InputConnect): Экземпляр класса InputConnect, содержащий требования к формированию класса DataSet

        Returns:
            list_vac (list[__main__.Vacancy]): Лист экземпляров Vacancy с данными о вакансиях
        """
        res_data = DataSet._сsv_reader(task.task_params['filename']['val'])
        ld = DataSet._csv_filer(res_data)
        list_vac = []
        for dct in ld:
            list_vac.append(Vacancy(dct))
        return list_vac

    def __init__(self, task):
        """Инициализирует экземпляр класса DataSet.

        Args:
            task (__main__.InputConnect): Экземпляр класса InputConnect, содержащий требования к формированию класса DataSet

        Returns:
            Экземпляр класса с заполненными свойствами vacancies_objects и dynamics_objects
        """
        self.file_name = task.task_params['filename']['val']
        self.vacancies_objects = DataSet._csv_parser(task)
        self.dynamics_objects = DynamicObjects(task, self.vacancies_objects)

class DynamicObjects:  
    """Класс для представления данных о всех вакансиях.

    Attributes:
        Свойства в виде словарей {наименование показателя: словарь вычисленных значений}
        salByYear (dict): Статистика для динамики уровня зарплат по годам 
        vacByYear (dict): Статистика для динамики количества вакансий по годам 
        salByYearProf (dict): Статистика для динамики уровня зарплат по годам для выбранной профессии
        vacByYearProf (dict): Статистика для динамики количества вакансий по годам для выбранной профессии 
        salByCity (dict): Статистика для уровеня зарплат по городам (в порядке убывания) 
        vacByCity (dict): Статистика для доли вакансий по городам (в порядке убывания)
    """
    def __init__(self, task, vacancies_objects):
        """Инициализирует экземпляр класса DynamicObjects.

        Args:
            task (__main__.InputConnect): Экземпляр класса InputConnect, содержащий требования к формированию класса DataSet
            vacancies_objects (list[Vacancy]): Лист экземпляров Vacancy хранящий данные о всех прочитанных из file_name вакансиях 

        Returns:
            Экземпляр класса с вычисленными значениями статистики для переданных task и vacancies_objects
        """
        self.salByYear     = {'name': 'Динамика уровня зарплат по годам', 'val': {}}                
        self.vacByYear     = {'name': 'Динамика количества вакансий по годам', 'val': {}} 
        self.salByYearProf = {'name': 'Динамика уровня зарплат по годам для выбранной профессии', 'val': {}}
        self.vacByYearProf = {'name': 'Динамика количества вакансий по годам для выбранной профессии', 'val': {}}
        self.salByCity     = {'name': 'Уровень зарплат по городам (в порядке убывания)', 'val': {}} 
        self.vacByCity     = {'name': 'Доля вакансий по городам (в порядке убывания)', 'val': {}}
        for vac in vacancies_objects:
            year = int(vac.published_at[0:4])
            sal_m = ((float(vac.salary.salary_to)+float(vac.salary.salary_from)) *
                    dic_money[vac.salary.salary_currency]['cost'])
            if (year in self.salByYear['val'].keys()):
                self.salByYear['val'][year] += sal_m
                self.vacByYear['val'][year] += 1
            else:
                self.salByYear['val'][year] = sal_m
                self.vacByYear['val'][year] = 1
            if (task.task_params['req_prof']['val'] in vac.name):
                if (year in self.salByYearProf['val'].keys()):
                    self.salByYearProf['val'][year] += sal_m
                    self.vacByYearProf['val'][year] += 1
                else:
                    self.salByYearProf['val'][year] = sal_m
                    self.vacByYearProf['val'][year] = 1
            city = vac.area_name
            if (city in self.salByCity['val'].keys()):
                self.salByCity['val'][city] += sal_m
                self.vacByCity['val'][city] += 1
            else:
                self.salByCity['val'][city] = sal_m
                self.vacByCity['val'][city] = 1
        for k in self.salByYear['val'].keys(): 
            self.salByYear['val'][k] = int(self.salByYear['val'][k] / (self.vacByYear['val'][k] * 2))
        self.salByYear['val'] = dict(sorted(self.salByYear['val'].items(), key = lambda x: x[0]))
        self.vacByYear['val'] = dict(sorted(self.vacByYear['val'].items(), key = lambda x: x[0]))
        if len(self.vacByYearProf['val']) == 0: 
            self.salByYearProf['val'] = {2022: 0}
            self.vacByYearProf['val'] = {2022: 0}
        else:
            for k in self.vacByYearProf['val'].keys():
                self.salByYearProf['val'][k] = int(self.salByYearProf['val'][k] / (self.vacByYearProf['val'][k] * 2))
        self.salByYearProf['val'] = dict(sorted(self.salByYearProf['val'].items(), key = lambda x: x[0]))
        self.vacByYearProf['val'] = dict(sorted(self.vacByYearProf['val'].items(), key = lambda x: x[0]))
        
        for c in self.salByCity['val'].keys():
            self.salByCity['val'][c] = int(self.salByCity['val'][c] / (self.vacByCity['val'][c] * 2))
        self.vacByCity['val'] = dict(filter(lambda x: x[1] >= len(vacancies_objects) / 100, self.vacByCity['val'].items()))
        self.salByCity['val'] = dict(filter(lambda x: self.vacByCity['val'].__contains__(x[0])  , self.salByCity['val'].items()))
        self.salByCity['val'] = dict(sorted(self.salByCity['val'].items(), key = lambda x: x[1], reverse=True))
        self.vacByCity['val'] = dict(sorted(self.vacByCity['val'].items(), key = lambda x: (-x[1])))
        for c in self.vacByCity['val']:
            self.vacByCity['val'][c] = round(self.vacByCity['val'][c] / len(vacancies_objects), 4)
        self.salByCity['val'] = dict(itertools.islice(self.salByCity['val'].items(), 10))
        self.vacByCity['val'] = dict(itertools.islice(self.vacByCity['val'].items(), 10))    
       
class InputConnect:
    """Класс, хранящий требования к обработке данных. 
       Инициализатор класса - по умолчанию - создает экземпляр с пустым значением 'val'.
       Значение по данному ключу заполняется при вызове метода get_task для экземпляра класса.

    Attributes: 
        task_params (dict): Словарь параметров обработки данных в виде {Название: результат ввода} 
    """
    task_params = {
    'filename': {'prompt': 'Введите название файла',     'val': ''},
    'req_prof': {'prompt': 'Введите название профессии', 'val': ''} 
    }

    @classmethod
    def get_task(self):
        """Метод, организующий диалог с пользователем и заполняющий свойства экземпляра класса.
           В методе осуществляется анализ вводимых данных и (возможно) требуемые заказчиком подстановки
           Т.к. количество используемых параметров сократилось до 2 - код сокращен
           В полном объеме код метода - копировать из 5.2 Функции в ООП

        Returns:
            Boolean: True - если успешно, False - если ошибка в запрашиваемых параметрах обработки
        """
        for tp in self.task_params.keys():
            self.task_params[tp]['val'] = ' '.join(input(f'{self.task_params[tp]["prompt"]}: ').split())
        return True
    
class Report:
    """Класс созданный по требованию заказчика для хранения и обработки данных из свойства dinamics_objects экземпляра класса DataSet

    Attributes: 
        salaries_year_level (dict):         то же что и dataset.dynamics_objects.salByYear['val']
        vacancies_year_count (dict):        то же что и dataset.dynamics_objects.vacByYear['val']
        selected_salary_year_level (dict):  то же что и dataset.dynamics_objects.salByYearProf['val']
        selected_vacancy_year_count (dict): то же что и dataset.dynamics_objects.vacByYearProf['val']
        salaries_city_level (dict):         то же что и dataset.dynamics_objects.salByCity['val']
        vacancies_city_count (dict):        то же что и dataset.dynamics_objects.vacByCity['val']
    """

    def __init__(self, dataset):
        """Инициализатор класса Report - просто копирует данные статистики в новую структуру

    Args:
        dataset (__main__.DataSet): экземпляр DataSet, содержащий в свойстве dynamics_objects всю необходимую статистику

    Returns:
        Заполненный экземпляр класса
        """
        self.salaries_year_level = dataset.dynamics_objects.salByYear['val']
        self.vacancies_year_count = dataset.dynamics_objects.vacByYear['val']
        self.selected_salary_year_level = dataset.dynamics_objects.salByYearProf['val']
        self.selected_vacancy_year_count = dataset.dynamics_objects.vacByYearProf['val']
        self.salaries_city_level = dataset.dynamics_objects.salByCity['val']
        self.vacancies_city_count = dataset.dynamics_objects.vacByCity['val']

    def generate_excel(self, req_prof):
        """Требуемый заказчиком метод генерации excel-файла. 
           Для генерации используются экземпляр класса Report и внешняя библиотека openpyxl
           Стилевое оформление таблиц устанавливается в вызываемом методе класса Report wb_style

        Args:
            req_prof (str): Наименование запрашиваемой профессии (используется для формирования имен столбцов в report.xlsx)
        
        Returns:
            Нет. Метод просто создает(перезаписывает) файл 'report.xlsx'
        """
        workbook = Workbook()
        stats_by_year = workbook.worksheets[0]
        stats_by_year.title = "Cтатистика по годам"
        stats_by_city = workbook.create_sheet("Cтатистика по городам")
        stats_by_year.append(["Год", "Средняя зарплата", f"Средняя зарплата - {req_prof}",
                              "Количество вакансий", f"Количество вакансий - {req_prof}"])
        for i, year in enumerate(self.salaries_year_level.keys(), 2):
            stats_by_year.cell(row=i, column=1, value=year)
            for j, dictionary in enumerate((self.salaries_year_level, self.vacancies_year_count,
                        self.selected_salary_year_level, self.selected_vacancy_year_count), 2):
                stats_by_year.cell(row=i, column=j, value=dictionary[year])
        stats_by_city.append(["Город", "Уровень зарплат", "", "Город", "Доля вакансий"])
        for i, city in enumerate(self.salaries_city_level.keys(), 2):
            stats_by_city.cell(row=i, column=1, value=city)
            stats_by_city.cell(row=i, column=2, value=self.salaries_city_level[city])
        for i, city in enumerate(self.vacancies_city_count.keys(), 2):
            stats_by_city.cell(row=i, column=4, value=city)
            stats_by_city.cell(row=i, column=5, value=self.vacancies_city_count[city]).number_format = '0.00%'
        self.wb_style(workbook)
        workbook.save('report.xlsx')

    @staticmethod
    def wb_style(wb):
        """Внутренний метод класса report. Устанавливае требуемые стили ячеек документа excel

        Args:
            wb (WorkBook): Формируемый документ excel

        Returns:
            Нет. Метод просто заполняет данные о требуемом стилевом оформлении в переданном wb
        """
        bold_font = Font(bold=True)
        thin = Side(border_style="thin", color="000000")
        outline = Border(top=thin, left=thin, right=thin, bottom=thin)
        for worksheet in wb.worksheets:
            for column_cells in worksheet.columns:
                length = max(len(str(cell.value) if cell.value is not None else "") for cell in column_cells)
                worksheet.column_dimensions[column_cells[0].column_letter].width = length + 3
            for cell in worksheet[1]:
                cell.font = bold_font
            for column in tuple(worksheet.columns):
                if column[1].value is None:
                    continue
                for cell in column:
                    cell.border = outline

    def generate_image(self, req_prof):
        """Требуемый заказчиком метод генерации png-рисунка. Для генерации используются внешняя библиотека matplotlib
           
        Args:
            req_prof (str): Наименование запрашиваемой профессии (используется для формирования подписей в graph.png)
        
        Returns:
            Нет. Метод просто создает(перезаписывает) файл 'graph.png'
        """
        fig1, ((f11, f12), (f21, f22)) = plt.subplots(2, 2, figsize=(12, 7.5), layout='constrained')
        self.generate_salByYear_graph(f11, req_prof)
        self.generate_vacByYear_graph(f12, req_prof)
        self.generate_salByCity_graph(f21)
        self.generate_vacByCity_graph(f22)
        plt.savefig('graph.png')

    def generate_salByYear_graph(self, f, req_prof):
        """Метод генерации части graph.png - рисунка(диаграммы) для "Уровеня зарплат по годам". 
           Для генерации используются экземпляр класса Report и внешние библиотеки matplotlib и numpy
           
        Args:
            f (matplotlib.axes._subplots.AxesSubplot): Подрисунок для graph.png
            req_prof (str): Наименование запрашиваемой профессии (используется для формирования подписей)
        
        Returns:
            Нет. Метод просто заполняет свойства переданной структуры подрисунка
        """
        f_labels = self.salaries_year_level.keys()
        x = np.arange(len(f_labels))
        width = 0.35
        f.bar(x - width / 2, self.salaries_year_level.values(), width, label='Средняя з/п')
        f.bar(x + width / 2, self.selected_salary_year_level.values(), width, label=f'З/п {req_prof}')
        f.set_xticks(x, f_labels, fontsize=8, rotation=90, ha='right')
        f.set_title("Уровень зарплат по годам")
        f.yaxis.grid(True)
        f.legend(fontsize=8, loc='upper left')

    def generate_vacByYear_graph(self, f, req_prof):
        """Метод генерации части graph.png - рисунка(диаграммы) для "Количества вакансий по годам". 
           Для генерации используются экземпляр класса Report и внешние библиотеки matplotlib и numpy
           
        Args:
            f (matplotlib.axes._subplots.AxesSubplot): Подрисунок для graph.png
            req_prof (str): Наименование запрашиваемой профессии (используется для формирования подписей)
        
        Returns:
            Нет. Метод просто заполняет свойства переданной структуры подрисунка
        """
        f_labels = self.vacancies_year_count.keys()
        x = np.arange(len(f_labels))
        width = 0.35
        f.bar(x - width / 2, self.vacancies_year_count.values(), width, label='Количество вакансий')
        f.bar(x + width / 2, self.selected_vacancy_year_count.values(), label=f'Количество вакансий {req_prof}')
        f.set_xticks(x, f_labels, fontsize=8, rotation=90, ha='right')
        f.set_title("Количество вакансий по годам")
        f.yaxis.grid(True)
        f.legend(fontsize=8, loc='upper left')

    def generate_salByCity_graph(self, f):
        """Метод генерации части graph.png - рисунка(диаграммы) для "Уровеня зарплат по городам". 
           Для генерации используются экземпляр класса Report и внешние библиотеки matplotlib и numpy
           
        Args:
            f (matplotlib.axes._subplots.AxesSubplot): Подрисунок для graph.png
                
        Returns:
            Нет. Метод просто заполняет свойства переданной структуры подрисунка
        """
        f_labels = self.salaries_city_level.keys()
        y_pos = np.arange(len(f_labels))
        f.barh(y_pos, self.salaries_city_level.values(), align='center')
        f.set_yticks(y_pos, fontsize=8, labels=f_labels)
        f.invert_yaxis()
        f.xaxis.grid(True)
        f.set_title("Уровень зарплат по городам")

    def generate_vacByCity_graph(self, f):
        """Метод генерации части graph.png - рисунка(круговой диаграммы) для "Доли вакансий по городам". 
           Для генерации используются экземпляр класса Report и внешние библиотеки matplotlib и numpy
           
        Args:
            f (matplotlib.axes._subplots.AxesSubplot): Подрисунок для graph.png
                
        Returns:
            Нет. Метод просто заполняет свойства переданной структуры подрисунка
        """
        f_labels = list(self.vacancies_city_count.keys())
        values = list(self.vacancies_city_count.values())
        f_labels.append('Другие')
        values.append(1 - sum(values))
        f.pie(values, labels=f_labels, textprops={'fontsize': 8}, startangle=0, labeldistance=1.1,
            colors=['tab:orange','tab:green','tab:red','tab:purple','tab:brown','tab:pink',
                    'tab:gray','tab:olive','tab:cyan','tab:blue','tab:blue'])
        f.set_title("Доля вакансий по городам")

    def generate_pdf(self, req_prof):
        """Требуемый заказчиком метод генерации pdf-файла. 
           Для генерации используются экземпляр класса Report, внешние библиотеки jinja2 и pdfkit, 
           сторонняя программа wkhtmltopdf.exe и шаблон генерируемого файла pdf_template.html
           
        Args:
            req_prof (str): Наименование запрашиваемой профессии (используется для формирования подписей)
        
        Returns:
            Нет. Метод просто создает(перезаписывает) файл 'report.pdf'
        """
     
        h1, h2, h3 = (["Год", "Средняя зарплата", f"Средняя зарплата - {req_prof}", "Количество вакансий",
            f"Количество вакансий - {req_prof}"], ["Город", "Уровень зарплат"], ["Город", "Доля вакансий"])
        r1 = list(map(lambda year: [year] + [dict[year] for dict in (self.salaries_year_level, self.vacancies_year_count,
            self.selected_salary_year_level, self.selected_vacancy_year_count)], self.salaries_year_level.keys()))
        r2 = list(map(lambda city: [city, self.salaries_city_level[city]], self.salaries_city_level.keys()))
        r3 = list(map(lambda city: [city, f'{round(self.vacancies_city_count[city]*100,2)}%'], self.vacancies_city_count.keys()))
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")
        pdf_template = template.render(graph_name='graph.png',req_prof=req_prof,h1=h1,h2=h2,h3=h3,r1=r1,r2=r2,r3=r3)
        config = pdfkit.configuration(wkhtmltopdf=
            r'D:/LIZOK/_Практика PY/Pyton 2 курс/Тема2_1 Библиотеки/02_01_03 PDF/wkhtmltopdf/wkhtmltopdf.exe')
        options = {'enable-local-file-access': None}
        pdfkit.from_string(pdf_template, 'report.pdf', options=options, configuration=config)

#####  Исполняемая часть кода  ######################################################################################

# Ввод данных пользователя:  
my_task = InputConnect()                        # Создаем пустой экземпляр InputConnect - my_task
my_task.get_task()                              # Заполняем поля этого экземпляра в диалоге с пользователем 
my_data = DataSet(my_task)                      # Создаем заполненный экземпляр DataSet - данные о вакансиях и статистика
                                                # в соответствии с запросами пользователя из my_task
if (my_data.vacancies_objects == None): exit()  # Если результатов нет - выходим

# Результаты есть - печатаем в требуемом виде
print(f'{my_data.dynamics_objects.salByYear["name"]}: {my_data.dynamics_objects.salByYear["val"]}')
print(f'{my_data.dynamics_objects.vacByYear["name"]}: {my_data.dynamics_objects.vacByYear["val"]}')
print(f'{my_data.dynamics_objects.salByYearProf["name"]}: {my_data.dynamics_objects.salByYearProf["val"]}')
print(f'{my_data.dynamics_objects.vacByYearProf["name"]}: {my_data.dynamics_objects.vacByYearProf["val"]}')
print(f'{my_data.dynamics_objects.salByCity["name"]}: {my_data.dynamics_objects.salByCity["val"]}')
print(f'{my_data.dynamics_objects.vacByCity["name"]}: {my_data.dynamics_objects.vacByCity["val"]}')

# Формируем экземпляр класса Report для имеющегося экземпляра DataSet - my_data
# Генерируем требуемые отчеты (report.xlsx, graph.png, report.pdf) для требуемой профессии
my_report = Report(my_data)
my_report.generate_excel(my_task.task_params['req_prof']['val'])
my_report.generate_image(my_task.task_params['req_prof']['val'])
my_report.generate_pdf(my_task.task_params['req_prof']['val'])